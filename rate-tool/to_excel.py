"""
to_excel.py
───────────
從 COSCO PDF 萃取費率資料並輸出成 Excel。

包含：
  - Direct Ports tab：原始萃取結果
  - Outports tab：原始萃取結果
  - All Rates tab：Direct Ports + Outports 合併，且 POL 用 '/' 拆開成獨立列

用法：python to_excel.py <path_to_pdf>
輸出：pdf_extract.xlsx（執行目錄）

依賴：pip install pdfplumber openpyxl
"""

import sys
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from coscopdf_extract import extract, split_pol, extract_rate_ref, extract_us_inland


HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def write_rate_sheet(ws, title, rows):
    ws.title = title
    for col, h in enumerate(["POL", "POD", "20' (USD)", "40'/HQ (USD)"], 1):
        cell = ws.cell(1, col, h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 20
    for r_idx, row in enumerate(rows, 2):
        for c_idx, key in enumerate(["pol", "pod", "rate_20", "rate_40"], 1):
            cell = ws.cell(r_idx, c_idx, row[key])
            cell.font      = DATA_FONT
            cell.alignment = LEFT if c_idx <= 2 else CENTER
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14




def write_ramp_sheet(ws, title, rows, rate_ref):
    ws.title = title

    ref_cell = ws.cell(1, 1, "Rate Ref:")
    ref_cell.font = HEADER_FONT
    val_cell = ws.cell(1, 2, rate_ref or "(not found)")
    val_cell.font = DATA_FONT

    for col, h in enumerate(["Location", "VIA POD", "20'box (USD)", "40'bx/HC (USD)"], 1):
        cell = ws.cell(3, col, h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
    ws.row_dimensions[3].height = 20
    for r_idx, row in enumerate(rows, 4):
        for c_idx, key in enumerate(["location", "via_pod", "rate_20", "rate_40"], 1):
            val = row[key]
            cell = ws.cell(r_idx, c_idx, val if val is not None else "-")
            cell.font      = DATA_FONT
            cell.alignment = LEFT if c_idx <= 2 else CENTER
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16


def main():
    if len(sys.argv) < 2:
        print("Usage: python to_excel.py <path_to_pdf>")
        sys.exit(1)

    pdf_path = sys.argv[1]
    direct, outports = extract(pdf_path)

    # 合併 + 拆開 POL
    combined = direct + outports
    all_rates = split_pol(combined)

    rate_ref  = extract_rate_ref(pdf_path)
    ramp_rows = extract_us_inland(pdf_path)

    wb = Workbook()
    write_rate_sheet(wb.active,         "Direct Ports", direct)
    write_rate_sheet(wb.create_sheet(), "Outports",     outports)
    write_rate_sheet(wb.create_sheet(), "All Rates",    all_rates)
    write_ramp_sheet(wb.create_sheet(), "US Inland",    ramp_rows, rate_ref)

    out_path = os.path.join(os.getcwd(), "pdf_extract.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
