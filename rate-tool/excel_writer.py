"""
excel_writer.py
───────────────
讀取 cheatsheet 的 Mapping tab，
把從 PDF 萃取的資料寫入 cheatsheet：
  - OFT sheet 的 20' / 40' / 40HC（依 POR + POD 對照）
  - OFT!B1 的 Rate Reference（例如 'TLI GL JULY'）
  - US inland sheet 的 20DV / 40DV/40HQ（依 Location 對照）

流程（OFT）：
1. 讀取 Mapping tab → 建立 (PDF_POL, PDF_POD) → (POR, POD) 對照字典
2. 掃描任意 row 找到 POD header → 得到 pod_col 和 header_row
3. POD 右邊 +2 格是 20'，往上驗證 header 是 "20'" → 確認欄位正確
4. +3 = 40'，+4 = 40HC
5. 掃描資料列，用 POR + POD 匹配，寫入 rate

流程（US inland）：
1. 找 Location / 20DV / 40DV/40HQ header
2. 用 Location 文字（大小寫不分）直接對照 PDF 萃取結果，寫入 rate
"""

import openpyxl


def _find_header(ws, keyword):
    """
    掃描整張 sheet，找第一個 value == keyword 的 cell。
    回傳 (row, col) 或拋出 ValueError。
    """
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or "").strip().upper() == keyword.upper():
                return cell.row, cell.column
    raise ValueError(f"找不到 header「{keyword}」")


def _find_oft_cols(ws, pod_col, header_row):
    """
    從 POD 欄往右 +2 格找 OFT 的 20' 欄，
    往上掃 header 驗證那格確實是 '20''。
    確認後 +3 = 40'，+4 = 40HC。

    pod_col    : POD 的 col index
    header_row : POD header 所在的 row

    回傳 (col_20, col_40, col_40hc) 或拋出 ValueError。
    """
    col_20_candidate = pod_col + 2

    # 往上掃這欄，看有沒有任何 row 的 header 是 "20'"
    confirmed = False
    for r in range(1, header_row + 1):
        val = str(ws.cell(r, col_20_candidate).value or "").strip()
        if val == "20'":
            confirmed = True
            break

    if not confirmed:
        raise ValueError(
            f"POD 右邊 +2 格（col {col_20_candidate}）往上找不到 \"20'\" header，"
            f"請確認 cheatsheet 欄位結構"
        )

    col_20   = col_20_candidate
    col_40   = col_20 + 1
    col_40hc = col_20 + 2

    return col_20, col_40, col_40hc


def _load_mapping(wb):
    """
    讀取 Mapping tab，回傳：
    {(pdf_pol_upper, pdf_pod_upper): (cheatsheet_por_upper, cheatsheet_pod_upper)}
    """
    if "Mapping" not in wb.sheetnames:
        raise ValueError("找不到 Mapping tab，請確認 cheatsheet 含有 Mapping 分頁")

    ws = wb["Mapping"]
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        pdf_pol, pdf_pod, cs_por, cs_pod = [str(v or "").strip() for v in row[:4]]
        mapping[(pdf_pol.upper(), pdf_pod.upper())] = (cs_por.upper(), cs_pod.upper())
    return mapping


def _build_rate_lookup(all_rates, mapping):
    """
    把 PDF 萃取的 all_rates（list of dict）
    透過 mapping 轉成 {(POR_upper, POD_upper): (rate_20, rate_40)} 字典。
    """
    lookup = {}
    for row in all_rates:
        pdf_pol = row["pol"].strip().upper()
        pdf_pod = row["pod"].strip().upper()
        key = (pdf_pol, pdf_pod)
        if key in mapping:
            por, pod = mapping[key]
            lookup[(por, pod)] = (row["rate_20"], row["rate_40"])
    return lookup


def _write_oft_rates(wb, all_rates, max_scan_rows=38):
    """
    把 all_rates 寫入 cheatsheet 的 OFT sheet。

    回傳 (updated_count, skipped_rows)
      updated_count : 成功寫入的列數
      skipped_rows  : 找不到對應 rate 的 (row_num, por, pod) list
    """
    ws = wb["OFT"]

    mapping = _load_mapping(wb)
    rate_lookup = _build_rate_lookup(all_rates, mapping)

    por_header_row, por_col = _find_header(ws, "POR")
    pod_header_row, pod_col = _find_header(ws, "POD")

    col_20, col_40, col_40hc = _find_oft_cols(ws, pod_col, pod_header_row)

    # 先清空所有 OFT 欄位的值（20' / 40' / 40HC）
    # 只清空純數值的格子，有公式的格子不動
    data_start_row = max(por_header_row, pod_header_row) + 1
    scan_end_row = data_start_row + max_scan_rows - 1
    for row_num in range(data_start_row, scan_end_row + 1):
        for col in (col_20, col_40, col_40hc):
            cell = ws.cell(row_num, col)
            if cell.data_type != 'f':   # 'f' = formula，公式格不動
                cell.value = None

    updated_count = 0
    skipped_rows = []

    for row_num in range(data_start_row, scan_end_row + 1):
        por = str(ws.cell(row_num, por_col).value or "").strip().upper()
        pod = str(ws.cell(row_num, pod_col).value or "").strip().upper()

        if not por or not pod:
            continue

        key = (por, pod)
        if key not in rate_lookup:
            skipped_rows.append((row_num, por, pod))
            continue

        rate_20, rate_40 = rate_lookup[key]
        ws.cell(row_num, col_20).value   = rate_20
        ws.cell(row_num, col_40).value   = rate_40
        ws.cell(row_num, col_40hc).value = rate_40   # 40HC = 40'

        updated_count += 1

    return updated_count, skipped_rows


def _write_rate_ref(wb, rate_ref):
    """
    把 Rate Reference（例如 'TLI GL JULY'）寫入 OFT sheet 的 B1。
    rate_ref 為 None 時不動作。
    """
    if not rate_ref:
        return False

    ws = wb["OFT"]
    cell = ws["B1"]
    if cell.data_type != 'f':   # 公式格不動
        cell.value = rate_ref
        return True
    return False


def _write_us_inland(wb, ramp_rows, max_scan_rows=40):
    """
    把 ramp_rows（extract_us_inland 產出）寫入 'US inland' sheet 的
    20DV / 40DV/40HQ 欄，用 Location 文字（大小寫不分）對照。

    回傳 (updated_count, skipped_rows, mismatched_rows)
      updated_count    : 成功寫入的列數（Location + Routing via 都相符）
      skipped_rows     : cheatsheet 裡找不到對應 Location 的 (row_num, location) list
      mismatched_rows  : Location 相符但 Routing via 對不上的
                          (row_num, location, cheatsheet_via_pod, pdf_via_pod) list
                          （不會寫入，需要人工確認）
    """
    if "US inland" not in wb.sheetnames:
        raise ValueError("找不到 'US inland' tab")

    ws = wb["US inland"]

    header_row, loc_col = _find_header(ws, "Location")
    _, via_col = _find_header(ws, "Routing via")
    _, col_20 = _find_header(ws, "20DV")
    _, col_40 = _find_header(ws, "40DV/40HQ")

    ramp_lookup = {r["location"].strip().upper(): r for r in ramp_rows}

    data_start_row = header_row + 1
    scan_end_row = data_start_row + max_scan_rows - 1

    updated_count = 0
    skipped_rows = []
    mismatched_rows = []

    for row_num in range(data_start_row, scan_end_row + 1):
        location = str(ws.cell(row_num, loc_col).value or "").strip().upper()
        if not location:
            continue

        if location not in ramp_lookup:
            skipped_rows.append((row_num, location))
            continue

        ramp = ramp_lookup[location]

        # Routing via 比對：PDF 的 via_pod 可能是合併寫法（例如 'ORF / NYC'），
        # 也可能帶尾綴字（例如 'NYC only'）。用 '/' 切開後只取每段的第一個
        # 詞（實際的港口代碼），忽略 'only' 這類描述字。
        cheatsheet_via = str(ws.cell(row_num, via_col).value or "").strip().upper()
        pdf_via_parts = [
            p.strip().split()[0].upper()
            for p in ramp["via_pod"].split("/")
            if p.strip()
        ]

        if cheatsheet_via and cheatsheet_via not in pdf_via_parts:
            mismatched_rows.append((row_num, location, cheatsheet_via, ramp["via_pod"]))
            continue

        cell_20 = ws.cell(row_num, col_20)
        if cell_20.data_type != 'f':
            cell_20.value = ramp["rate_20"] if ramp["rate_20"] is not None else "-"

        cell_40 = ws.cell(row_num, col_40)
        if cell_40.data_type != 'f':
            cell_40.value = ramp["rate_40"] if ramp["rate_40"] is not None else "-"

        updated_count += 1

    return updated_count, skipped_rows, mismatched_rows


def update_cheatsheet(excel_path, all_rates, rate_ref=None, ramp_rows=None,
                       output_path=None, max_scan_rows=38):
    """
    主函式：把 PDF 萃取結果寫入 cheatsheet（單次讀檔、單次存檔）。

    參數：
      excel_path  : cheatsheet 的路徑（含 Mapping tab）
      all_rates   : coscopdf_extract.extract() + split_pol() 產出的 list of dict
                    每筆含 pol / pod / rate_20 / rate_40
      rate_ref    : coscopdf_extract.extract_rate_ref() 產出的字串，例如 'TLI GL JULY'
                    None 則不寫入 B1
      ramp_rows   : coscopdf_extract.extract_us_inland() 產出的 list of dict
                    None 或空 list 則不更新 US inland tab
      output_path : 輸出路徑，None 則直接覆蓋原檔

    回傳 dict：
      {
        "oft_updated":      int,
        "oft_skipped":      [(row_num, por, pod), ...],
        "rate_ref_written": bool,
        "inland_updated":   int,
        "inland_skipped":   [(row_num, location), ...],
        "inland_mismatched": [(row_num, location, cheatsheet_via, pdf_via), ...],
      }
    """
    wb = openpyxl.load_workbook(excel_path)

    oft_updated, oft_skipped = _write_oft_rates(wb, all_rates, max_scan_rows=max_scan_rows)
    rate_ref_written = _write_rate_ref(wb, rate_ref)

    inland_updated, inland_skipped, inland_mismatched = 0, [], []
    if ramp_rows:
        inland_updated, inland_skipped, inland_mismatched = _write_us_inland(wb, ramp_rows)

    out = output_path or excel_path
    wb.save(out)

    return {
        "oft_updated":       oft_updated,
        "oft_skipped":       oft_skipped,
        "rate_ref_written":  rate_ref_written,
        "inland_updated":    inland_updated,
        "inland_skipped":    inland_skipped,
        "inland_mismatched": inland_mismatched,
    }


def update_oft_rates(excel_path, all_rates, output_path=None, max_scan_rows=38):
    """
    向後相容用：只更新 OFT sheet（不動 B1 / US inland）。
    回傳 (updated_count, skipped_rows)。
    """
    wb = openpyxl.load_workbook(excel_path)
    updated_count, skipped_rows = _write_oft_rates(wb, all_rates, max_scan_rows=max_scan_rows)
    out = output_path or excel_path
    wb.save(out)
    return updated_count, skipped_rows
